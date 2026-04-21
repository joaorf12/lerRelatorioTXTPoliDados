import pandas as pd
import re
import unicodedata
from datetime import datetime
from openpyxl import load_workbook
import os

# ================================
# 1. FUNÇÕES DE APOIO (TRATAMENTO)
# ================================

def normalizar_tipo(tipo):
    if pd.isna(tipo): return tipo
    tipo = str(tipo).upper().strip()
    
    # Remove acentos
    tipo = ''.join(c for c in unicodedata.normalize('NFD', tipo) if unicodedata.category(c) != 'Mn')
    
    substituicoes = {
        "BARRAS": "BARRA",
        "GRADES": "GRADE",
        "PAINEIS": "PAINEL",
        "PES": "PÉS",
        "LAQ": "LAQUEADO",
        "KDM": "",
        "BRANCA": "BRANCO",
        "ESCADA": "ESC"  # Transforma ESCADA em ESC
    }
    for busca, substituto in substituicoes.items():
        tipo = tipo.replace(busca, substituto).strip()
    return tipo

def extrair_modelo_avancado(desc):
    # Remove todos os espaços para não confundir "B 105" com "B105"
    desc_limpa = str(desc).upper().replace(" ", "")
    
    # Modelos Base (A ordem de busca é do maior para o menor)
    if "B105" in desc_limpa: return "B105"
    if "B104" in desc_limpa: return "B104"
    if "B142" in desc_limpa: return "B142"
    if "B039" in desc_limpa: return "B039"
    
    if "O0023" in desc_limpa: return "O00230"
    if "O0024" in desc_limpa: return "O00246"
    if "O0103" in desc_limpa: return "O01030"
    if "O0104" in desc_limpa: return "O01046"
    
    if "C7233" in desc_limpa: return "C72330"
    if "C7234" in desc_limpa: return "C72346"
    
    if "O005" in desc_limpa: return "O005"
    if "O016" in desc_limpa: return "O016"
    
    return None

# ================================
# 2. LER E PROCESSAR TXT
# ================================
data_hoje = datetime.now().strftime("%d_%m_%Y")
dados = []
with open(f"estoque{data_hoje}.txt", "r", encoding="cp1252") as f:
    for linha in f:
        if not re.search(r'\b\d{5}\b', linha):
            continue
        
        codigo = re.search(r'\b\d{5}\b', linha)
        desc_match = re.search(r'\d{5}\s+(.*?)\s+\d+,\d+', linha)
        numeros = re.findall(r'\d+,\d+', linha)
        
        if codigo and desc_match and numeros:
            descricao = desc_match.group(1).strip()
            estoque = float(numeros[-1].replace(',', '.'))
            
            dados.append({
                "CODIGO": codigo.group(),
                "DESCRICAO": descricao,
                "ESTOQUE": estoque
            })

df_rel = pd.DataFrame(dados)
df_rel["DESCRICAO"] = df_rel["DESCRICAO"].astype(str)

# Extração de Volume e Cor do TXT
df_rel[["DESCRICAO_LIMPA", "VOLUME_STR"]] = df_rel["DESCRICAO"].str.extract(r'^(.*?)\s+(VOLUME(?:\s+\w+)*)$', flags=re.IGNORECASE).fillna("")
df_rel["COR"] = df_rel["DESCRICAO_LIMPA"].str.extract(r'(BRANCO|CAST[^\s]*|MEL)', flags=re.IGNORECASE)[0].str.upper()

# Criar DataFrame Final de Trabalho
df_final = pd.DataFrame()
df_final["DESCRICAO"] = df_rel["DESCRICAO_LIMPA"]
df_final["COR"] = df_rel["COR"]
df_final["VOLUME_ORIGINAL"] = df_rel["VOLUME_STR"]
df_final["ESTOQUE_QUANTIDADE"] = df_rel["ESTOQUE"]

# Tratamento de Tipo e Modelo
df_final["TIPO_FINAL"] = df_final["VOLUME_ORIGINAL"].str.replace("VOLUME", "", case=False).str.strip()
df_final["TIPO_FINAL"] = df_final["TIPO_FINAL"].apply(normalizar_tipo)

# Ajuste inteligente para Escadas (adiciona o 01 ou 02 no tipo se existir na descrição)
def ajustar_tipo_escada(row):
    tipo = row["TIPO_FINAL"]
    desc = str(row["DESCRICAO"]).upper()
    if tipo == "ESC":
        if " 01 " in desc or "01 E" in desc: return "ESC 01"
        if " 02 " in desc or "02 E" in desc: return "ESC 02"
    return tipo

df_final["TIPO_FINAL"] = df_final.apply(ajustar_tipo_escada, axis=1)
df_final["MODELO"] = df_final["DESCRICAO"].apply(extrair_modelo_avancado)

# Sigla da Cor e Nome do Produto
mapa_cor = {"BRANCO": "BR", "CAST": "CC", "MEL": "ML"}
df_final["COR_SIGLA"] = df_final["COR"].map(mapa_cor)
df_final["PRODUTO"] = df_final["MODELO"].fillna("") + df_final["COR_SIGLA"].fillna("")

# ================================
# 3. MAPA DE TIPOS (TXT -> EXCEL)
# ================================
mapa_tipo = {
    "B105": {"GRADE CASAL": "V1", "GRADE SOLTEIRO": "V2", "GRADE": "V1", "PÉS": "V3", "BARRA": "V4", "ESC 01": "V5 01", "ESC 02": "V5 02"},
    "B104": {"GRADE": "V1", "BARRA": "V2", "ESC 01": "V3 01", "ESC 02": "V3 02"},
    "B142": {"GRADE SOLTEIRO": "V1", "GRADE CASAL": "V2", "GRADE": "V1", "PÉS": "V3", "BARRA": "V4", "ESC 01": "V5 01", "ESC 02": "V5 02"},
    "B039": {"PAINEL": "V1", "GRADE": "V1", "PÉS": "V2", "BARRA": "V3", "ESC 01": "V4 01", "ESC 02": "V4 02"},
    "C72330": {"PAINEL": "V1", "PÉS": "V2", "BARRA": "V3"},
    "C72346": {"PAINEL": "V1", "PÉS": "V2", "BARRA": "V3"},
    "O005": {"CABECEIRA": "V1", "BARRA": "V2"},
    "O016": {"CABECEIRA": "V1", "BARRA": "V2", "GRADE": "V1"},
    "O00230": {"": "V1"}, "O01030": {"V1": "V1", "V2": "V2"},
    "O00246": {"": "V1"}, "O01046": {"V1": "V1", "V2": "V2"},
}

def mapear_coluna(row):
    modelo, tipo = row["MODELO"], row["TIPO_FINAL"]
    if modelo in mapa_tipo:
        mapa = mapa_tipo[modelo]
        if tipo in mapa: return mapa[tipo]
        if str(tipo).isdigit(): return f"V{tipo}"
    return None

df_final["COLUNA_DESTINO"] = df_final.apply(mapear_coluna, axis=1)

# Debug de falhas
nao_mapeados = df_final[df_final["COLUNA_DESTINO"].isna()]
if not nao_mapeados.empty:
    print("\n⚠️ NÃO MAPEADOS (Verifique se o tipo existe no mapa_tipo):")
    print(nao_mapeados[["DESCRICAO", "TIPO_FINAL", "MODELO"]].drop_duplicates())
else:
    print("\n✅ Todos os itens mapeados com sucesso!")

# ================================
# 4. PROCESSAR BLOCOS DA PLANILHA
# ================================
df_raw = pd.read_excel("contagem produtos finalizados.xlsx", header=None)

blocos = {
    "B105": {"linha_volumes": 2, "linha_tipos": 3, "inicio_dados": 4, "fim_dados": 8},
    "B104": {"linha_volumes": 9, "linha_tipos": 10, "inicio_dados": 11, "fim_dados": 15},
    "O00230": {"linha_volumes": 16, "linha_tipos": 17, "inicio_dados": 18, "fim_dados": 22},
    "O01030": {"linha_volumes": 23, "linha_tipos": 24, "inicio_dados": 25, "fim_dados": 28},
    "O00246": {"linha_volumes": 29, "linha_tipos": 30, "inicio_dados": 31, "fim_dados": 35},
    "O01046": {"linha_volumes": 36, "linha_tipos": 37, "inicio_dados": 38, "fim_dados": 41},
    "C72330": {"linha_volumes": 42, "linha_tipos": 43, "inicio_dados": 44, "fim_dados": 46},
    "C72346": {"linha_volumes": 42, "linha_tipos": 43, "inicio_dados": 48, "fim_dados": 50},
    "O005": {"linha_volumes": 51, "linha_tipos": 52, "inicio_dados": 53, "fim_dados": 57},
    "O016": {"linha_volumes": 51, "linha_tipos": 52, "inicio_dados": 59, "fim_dados": 61},
    "B142": {"linha_volumes": 62, "linha_tipos": 63, "inicio_dados": 64, "fim_dados": 68},
    "B039": {"linha_volumes": 69, "linha_tipos": 70, "inicio_dados": 71, "fim_dados": 75},
}

# ================================
# 5. GRAVAR NO EXCEL MANTENDO ESTILO
# ================================
wb = load_workbook("contagem produtos finalizados.xlsx")
ws = wb.active

for modelo, config in blocos.items():
    linha_vols = df_raw.iloc[config["linha_volumes"]]
    
    for i in range(config["inicio_dados"], config["fim_dados"] + 1):
        produto_planilha = str(df_raw.iloc[i, 1]).strip()
        
        for col in range(len(df_raw.columns)):
            volume_planilha = str(linha_vols[col]).strip()
            
            match = df_final[
                (df_final["PRODUTO"] == produto_planilha) & 
                (df_final["COLUNA_DESTINO"] == volume_planilha)
            ]
            
            if not match.empty:
                novo_valor = match["ESTOQUE_QUANTIDADE"].values[0]
                ws.cell(row=i+1, column=col+1, value=novo_valor)

# Salvar arquivo
nome_arquivo = f"resultado_{data_hoje}.xlsx"

os.makedirs("resultado", exist_ok=True)
caminho_arquivo = os.path.join("resultado", nome_arquivo)

wb.save(caminho_arquivo)

print(f"\n✅ Sucesso! Planilha gerada: {caminho_arquivo}")