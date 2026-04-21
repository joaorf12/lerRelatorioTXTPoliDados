import pandas as pd
import re
import unicodedata
from datetime import datetime
from openpyxl import load_workbook
import os

# ================================
# 1. FUNÇÕES DE APOIO (TRATAMENTO)
# ================================

def normalizar_tipo(tipo):
    if pd.isna(tipo) or tipo is None: 
        return ""
    
    # Converte para string e limpa espaços
    tipo = str(tipo).upper().strip()

    nfkd = unicodedata.normalize('NFKD', str(tipo))
    
    # Remove acentos (Trata CASTÃNHO para CASTANHO, PÉS para PES antes da volta, etc)
    tipo = "".join([c for c in nfkd if not unicodedata.category(c).startswith('M')]).upper()
    
    # Dicionário de padronização
    substituicoes = {
        "BARRAS": "BARRA",
        "GRADES": "GRADE",
        "PAINEIS": "PAINEL",
        "PES": "PÉS",
        "LAQ": "LAQUEADO",
        "KDM": "",
        "BRANCA": "BRANCO",
        "CASTANHO": "CASTANHO",
        "ESCADA": "ESC"
    }
    
    # Aplica as substituições apenas se a palavra inteira estiver lá
    palavras = tipo.split()
    novas_palavras = []
    for p in palavras:
        # Verifica se a palavra está no dicionário, se não, mantém a original
        novas_palavras.append(substituicoes.get(p, p))
    
    resultado = " ".join(novas_palavras).strip()
    
    # Remove espaços duplos
    return re.sub(r'\s+', ' ', resultado)

def extrair_modelo_avancado(desc):
    # Remove todos os espaços para não confundir "B 105" com "B105"
    desc_limpa = str(desc).upper().replace(" ", "")
    
    # Modelos Base (ordem: do mais específico para o mais genérico)
    if "B105" in desc_limpa: return "B105"
    if "B104" in desc_limpa: return "B104"
    if "B142" in desc_limpa: return "B142"
    if "B039" in desc_limpa: return "B039"
    
    if "O0023" in desc_limpa: return "O00230"
    if "O0024" in desc_limpa: return "O00246"
    if "O0103" in desc_limpa: return "O01030"
    if "O0104" in desc_limpa: return "O01046"
    
    if "C7233" in desc_limpa: return "C72330"
    if "C7234" in desc_limpa: return "C72346"
    
    if "O005" in desc_limpa: return "O005"
    if "O016" in desc_limpa: return "O016"
    
    return None

def extrair_cor_sigla(cor_str):
    """Mapeia a cor extraída para sigla, tratando variações como CASTANHO, CASTANHO CLARO etc."""
    if pd.isna(cor_str):
        return None
    
    cor = str(cor_str).upper().strip()

    if cor == "BRANCA/CAST":
        return "BR/CC"
    if cor.startswith("CAST"):
        return "CC"
    if cor.startswith("BRANC"):
        return "BR"
    if cor == "MEL":
        return "ML"
    return None

def tratar_nome_cor(x):
    nome = str(x).upper()
    # 1º: Verifica o caso mais específico (Combo)
    if "BRANC" in nome and "CAST" in nome:
        return "BRANCA/CASTANHO"
    # 2º: Verifica o caso genérico
    if nome.startswith("BRANC"):
        return "BRANCO"
    return x

# ================================
# 2. LER E PROCESSAR TXT
# ================================
data_hoje = datetime.now().strftime("%d_%m_%Y")
dados = []
with open(f"estoque{data_hoje}.txt", "r", encoding="cp1252") as f:
    for linha in f:
        if not re.search(r'\b\d{5}\b', linha):
            continue
        
        codigo = re.search(r'\b\d{5}\b', linha)
        desc_match = re.search(r'\d{5}\s+(.*?)\s+\d+,\d+', linha)
        numeros = re.findall(r'-?\d+(?:,\d+)?', linha)
        
        if codigo and desc_match and numeros:
            descricao = desc_match.group(1).strip()
            estoque = float(numeros[-1].replace(',', '.'))
            
            dados.append({
                "CODIGO": codigo.group(),
                "DESCRICAO": descricao,
                "ESTOQUE": estoque
            })

df_rel = pd.DataFrame(dados)
df_rel["DESCRICAO"] = df_rel["DESCRICAO"].astype(str)

# Extração de Volume e Cor do TXT
df_rel[["DESCRICAO_LIMPA", "VOLUME_STR"]] = df_rel["DESCRICAO"].str.extract(r'^(.*?)\s+(VOLUME(?:\s+\w+)*)$', flags=re.IGNORECASE).fillna("")
# 1. Ajuste a Regex para capturar BRANCA (Adicionado [AO] no final)
# A ordem aqui importa: colocamos o padrão mais específico primeiro
df_rel["COR"] = df_rel["DESCRICAO"].str.extract(
    r'(BRANC[AO]/CAST[^\s]*|BRANC[AO]|CAST[^\s]*|MEL)', 
    flags=re.IGNORECASE
)[0].str.upper()
# Criar DataFrame Final de Trabalho
df_final = pd.DataFrame()
df_final["DESCRICAO"] = df_rel["DESCRICAO_LIMPA"]
df_final["COR"] = df_rel["COR"]
df_final["VOLUME_ORIGINAL"] = df_rel["VOLUME_STR"]
df_final["ESTOQUE_QUANTIDADE"] = df_rel["ESTOQUE"]

# Tratamento de Tipo e Modelo
df_final["TIPO_FINAL"] = df_final["VOLUME_ORIGINAL"].str.replace("VOLUME", "", case=False).str.strip()
df_final["TIPO_FINAL"] = df_final["TIPO_FINAL"].apply(normalizar_tipo)

# Ajuste inteligente para Escadas (adiciona o 01 ou 02 no tipo se existir na descrição)
def ajustar_tipo_escada(row):
    tipo = row["TIPO_FINAL"]
    desc = str(row["DESCRICAO"]).upper()
    if tipo == "ESC":
        if " 01 " in desc or "01 E" in desc: return "ESC 01"
        if " 02 " in desc or "02 E" in desc: return "ESC 02"
    return tipo

df_final["TIPO_FINAL"] = df_final.apply(ajustar_tipo_escada, axis=1)
df_final["MODELO"] = df_final["DESCRICAO"].apply(extrair_modelo_avancado)

# Sigla da Cor — usando função robusta que trata CASTANHO, CASTANHO CLARO etc.
df_final["COR_SIGLA"] = df_final["COR"].apply(extrair_cor_sigla)

df_final["PRODUTO"] = df_final["MODELO"].fillna("") + df_final["COR_SIGLA"].fillna("")

# Ajuste final para o nome da cor no Debug aparecer como BRANCO e não vazio
df_final["COR"] = df_final["COR"].apply(tratar_nome_cor)

# ================================
# 3. MAPA DE TIPOS (TXT -> EXCEL)
# ================================
mapa_tipo = {
    "B105": {"GRADE CASAL": "V1", "GRADE SOLTEIRO": "V2", "GRADE": "V1", "PÉS": "V3", "BARRA": "V4", "ESC 01": "V5 01", "ESC 02": "V5 02"},
    "B104": {"GRADE": "V1", "BARRA": "V2", "ESC 01": "V3 01", "ESC 02": "V3 02"},
    "B142": {"GRADE CASAL": "V1", "GRADE SOLTEIRO": "V2", "GRADE": "V1", "PÉS": "V3", "BARRA": "V4", "ESC 01": "V5 01", "ESC 02": "V5 02"},
    "B039": {"PAINEL": "V1", "GRADE": "V1", "PÉS": "V2", "BARRA": "V3", "ESC 01": "V4 01", "ESC 02": "V4 02"},
    "C72330": {"PAINEL": "V1", "PÉS": "V2", "BARRA": "V3"},
    "C72346": {"PAINEL": "V1", "PÉS": "V2", "BARRA": "V3"},
    "O005": {"CABECEIRA": "V1", "BARRA": "V2"},
    "O016": {"CABECEIRA": "V1", "BARRA": "V2", "GRADE": "V1"},
    "O00230": {"": "V1"}, "O01030": {"V1": "V1", "V2": "V2"},
    "O00246": {"": "V1"}, "O01046": {"V1": "V1", "V2": "V2"},
}

def mapear_coluna(row):
    modelo, tipo = row["MODELO"], row["TIPO_FINAL"]
    if modelo in mapa_tipo:
        mapa = mapa_tipo[modelo]
        if tipo in mapa: return mapa[tipo]
        if str(tipo).isdigit(): return f"V{tipo}"
    return None

df_final["COLUNA_DESTINO"] = df_final.apply(mapear_coluna, axis=1)

# Debug de falhas
nao_mapeados = df_final[df_final["COLUNA_DESTINO"].isna()]
if not nao_mapeados.empty:
    print("\n⚠️ NÃO MAPEADOS (Verifique se o tipo existe no mapa_tipo):")
    print(nao_mapeados[["DESCRICAO", "TIPO_FINAL", "MODELO"]].drop_duplicates())
else:
    print("\n✅ Todos os itens mapeados com sucesso!")

# ========================================================
# SINCRONIZAÇÃO MANUAL DE ESCADAS (B105 -> B142)
# ========================================================
# 1. Filtramos apenas as linhas de escada do B105
escadas_b105 = df_final[
    (df_final["MODELO"] == "B105") & 
    (df_final["TIPO_FINAL"].str.contains("ESC", na=False))
].copy()

if not escadas_b105.empty:
    # 2. Mudamos o modelo e o nome do produto para B142
    escadas_b105["MODELO"] = "B142"
    # Recriamos o nome do produto para ficar B142BR, B142CC, etc.
    escadas_b105["PRODUTO"] = escadas_b105["MODELO"] + escadas_b105["COR_SIGLA"].fillna("")
    
    # 3. Adicionamos essas novas linhas ao DataFrame principal
    df_final = pd.concat([df_final, escadas_b105], ignore_index=True)
    #print(f"✅ Sincronização: {len(escadas_b105)} volumes de escada espelhados para o B142.")

# ========================================================

# ================================
# DEBUG: GERAR TXT DE COMPARAÇÃO
# ================================
os.makedirs("resultado", exist_ok=True)
caminho_debug = os.path.join("resultado", f"debug_comparacao_{data_hoje}.txt")

with open(caminho_debug, "w", encoding="utf-8") as f:
    f.write("=" * 90 + "\n")
    f.write("RELATORIO DE COMPARACAO — O QUE O SCRIPT LEU DO TXT\n")
    f.write(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
    f.write("=" * 90 + "\n")

    modelos_ordem = ["B105", "B104", "B142", "B039",
                     "O00230", "O00246", "O01030", "O01046",
                     "C72330", "C72346", "O005", "O016"]

    for modelo in modelos_ordem:
        grupo = df_final[df_final["MODELO"] == modelo].copy()
        if grupo.empty:
            continue
        f.write(f"\n{'─' * 90}\n")
        f.write(f"  MODELO: {modelo}\n")
        f.write(f"{'─' * 90}\n")
        f.write(f"  {'PRODUTO':<20} {'COR':<20} {'TIPO (VOLUME)':<22} {'COL.DESTINO':<14} {'QUANTIDADE':>10}\n")
        f.write(f"  {'─'*20} {'─'*20} {'─'*22} {'─'*14} {'─'*10}\n")
        for _, row in grupo.sort_values(["PRODUTO", "TIPO_FINAL"]).iterrows():
            produto  = str(row["PRODUTO"])
            cor      = str(row["COR"]) if pd.notna(row["COR"]) else "?"
            tipo     = str(row["TIPO_FINAL"]) if pd.notna(row["TIPO_FINAL"]) else "?"
            col_dest = str(row["COLUNA_DESTINO"]) if pd.notna(row["COLUNA_DESTINO"]) else "NAO MAPEADO"
            qtd      = row["ESTOQUE_QUANTIDADE"]
            f.write(f"  {produto:<20} {cor:<20} {tipo:<22} {col_dest:<14} {qtd:>10.0f}\n")

    sem_modelo = df_final[df_final["MODELO"].isna()]
    if not sem_modelo.empty:
        f.write(f"\n{'─' * 90}\n")
        f.write("  ITENS SEM MODELO IDENTIFICADO\n")
        f.write(f"{'─' * 90}\n")
        for _, row in sem_modelo.iterrows():
            f.write(f"  DESC: {row['DESCRICAO']}  |  TIPO: {row['TIPO_FINAL']}  |  QTD: {row['ESTOQUE_QUANTIDADE']:.0f}\n")

    f.write(f"\n{'=' * 90}\n")
    f.write(f"  TOTAL LIDO: {len(df_final)} itens  |  "
            f"MAPEADOS: {df_final['COLUNA_DESTINO'].notna().sum()}  |  "
            f"NAO MAPEADOS: {df_final['COLUNA_DESTINO'].isna().sum()}\n")
    f.write("=" * 90 + "\n")

# ================================
# 4. PROCESSAR BLOCOS DA PLANILHA
# ================================
df_raw = pd.read_excel("contagem produtos finalizados.xlsx", header=None)

blocos = {
    "B105": {"linha_volumes": 2, "linha_tipos": 3, "inicio_dados": 4, "fim_dados": 8},
    "B104": {"linha_volumes": 9, "linha_tipos": 10, "inicio_dados": 11, "fim_dados": 15},
    "O00230": {"linha_volumes": 16, "linha_tipos": 17, "inicio_dados": 18, "fim_dados": 22},
    "O01030": {"linha_volumes": 23, "linha_tipos": 23, "inicio_dados": 24, "fim_dados": 28},
    "O00246": {"linha_volumes": 29, "linha_tipos": 30, "inicio_dados": 31, "fim_dados": 35},
    "O01046": {"linha_volumes": 36, "linha_tipos": 36, "inicio_dados": 37, "fim_dados": 41},
    "C72330": {"linha_volumes": 42, "linha_tipos": 43, "inicio_dados": 44, "fim_dados": 46},
    "C72346": {"linha_volumes": 42, "linha_tipos": 43, "inicio_dados": 48, "fim_dados": 50},
    "O005": {"linha_volumes": 51, "linha_tipos": 52, "inicio_dados": 53, "fim_dados": 57},
    "O016": {"linha_volumes": 51, "linha_tipos": 52, "inicio_dados": 59, "fim_dados": 61},
    "B142": {"linha_volumes": 62, "linha_tipos": 63, "inicio_dados": 64, "fim_dados": 68},
    "B039": {"linha_volumes": 69, "linha_tipos": 70, "inicio_dados": 71, "fim_dados": 75},
}

# ================================
# 5. GRAVAR NO EXCEL MANTENDO ESTILO
# ================================
wb = load_workbook("contagem produtos finalizados.xlsx")
ws = wb.active

# Mostrar todas as colunas
pd.set_option('display.max_columns', None)

# Mostrar todas as linhas (cuidado se o DF for gigante!)
pd.set_option('display.max_rows', None)

for modelo, config in blocos.items():
    linha_vols = df_raw.iloc[config["linha_volumes"]]

    print(f"\n>>> Processando Bloco: {modelo}") # Debug de bloco
    
    for i in range(config["inicio_dados"], config["fim_dados"] + 1):
        produto_planilha = str(df_raw.iloc[i, 1]).strip()
        
        for col in range(len(df_raw.columns)):
            volume_planilha = str(linha_vols[col]).strip()

            # IGNORA COLUNAS VAZIAS OU INVÁLIDAS
            if volume_planilha in ['nan', '', 'None', 'TOTAL']:
                continue
            
            # Dentro do loop, ajuste a busca do match:
            produto_limpo = produto_planilha.replace(" ", "").upper()

            match = df_final[
                (df_final["PRODUTO"].str.replace(" ", "") == produto_limpo) & 
                (df_final["COLUNA_DESTINO"] == volume_planilha)
            ]
            
                        # Se falhar a busca exata, tenta uma busca parcial para o BRANCO
            if match.empty and "BR" in produto_limpo:
                match = df_final[
                    (df_final["PRODUTO"].str.contains("BRANCO|BR", regex=True)) & 
                    (df_final["PRODUTO"].str.contains(modelo)) &
                    (df_final["COLUNA_DESTINO"] == volume_planilha)
                ]

            if not match.empty:
                valor_estoque = float(match["ESTOQUE_QUANTIDADE"].values[0])
                print(f"GRAVANDO: {produto_planilha} em L:{i+1} C:{col+1} VALOR:{valor_estoque}")
                ws.cell(row=i+1, column=col+1, value=valor_estoque)
            else:
                # DEBUG CRÍTICO: Se for um produto que deveria ter estoque, me diga por que falhou
                if "O010" in produto_planilha or "O016" in produto_planilha:
                    print(f"DEBUG BRANCO: Tentou '{produto_planilha}' com Vol '{volume_planilha}' mas não achou no DF_FINAL")

# Salvar arquivo
nome_arquivo = f"resultado_{data_hoje}.xlsx"

os.makedirs("resultado", exist_ok=True)
caminho_arquivo = os.path.join("resultado", nome_arquivo)

wb.save(caminho_arquivo)

print(f"\n✅ Sucesso! Planilha gerada: {caminho_arquivo}")